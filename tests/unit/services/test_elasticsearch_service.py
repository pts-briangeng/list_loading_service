import copy
import csv
import httplib
import json
import os
import types
import unittest
import uuid

import mock
import openpyxl
from elasticsearch import helpers, exceptions
from liblcp import context
from nose import tools

import configuration
from app import models, services
from app.services import elastic
from tests import builders


class CsvMock(list):
    line_num = 1L


class MockCell(object):

    def __init__(self, cell_value):
        self.cell_value = cell_value

    @property
    def value(self):
        return self.cell_value


class MockHttpResponse(object):

    def __init__(self, status_code=httplib.OK, response=None):
        self.status_code = status_code
        self.response = response

    def json(self):
        return self.response

    def status_code(self):
        return self.status_code

    def raise_for_status(self):
        if self.status_code not in [httplib.OK, httplib.CREATED]:
            raise Exception


CORRELATION_ID = str(uuid.uuid4())
PRINCIPAL = str(uuid.uuid4())
context.set_headers_getter(lambda name: {context.HEADERS_EXTERNAL_BASE_URL: 'http://localhost',
                                         context.HEADERS_CORRELATION_ID: CORRELATION_ID,
                                         context.HEADERS_MODE: context.MODE_LIVE,
                                         context.HEADERS_PRINCIPAL: PRINCIPAL}[name])

NOT_FOUND_EXCEPTION = exceptions.TransportError(httplib.NOT_FOUND,
                                                'IndexMissingException[[123] missing]',
                                                {
                                                    'status': httplib.BAD_REQUEST,
                                                    'error': 'IndexMissingException[[123] missing]'
                                                })

INTERNAL_SERVER_ERROR_EXCEPTION = exceptions.TransportError(httplib.INTERNAL_SERVER_ERROR,
                                                            'Internal Server error',
                                                            {
                                                                'status': httplib.INTERNAL_SERVER_ERROR,
                                                                'error': 'Server error'
                                                            })


class TestElasticSearchService(unittest.TestCase):

    def setUp(self):
        self.data = {
            'url': 'url',
            'filePath': 'file.csv',
            'service': 'service',
            'list_id': 'id',
            'callbackUrl': 'callback',
        }
        self.member_data = copy.deepcopy(self.data)
        self.member_data['member_id'] = 'member_id'
        self.service = services.ElasticSearch()
        configuration.configure_from(os.path.join(configuration.CONFIGURATION_PATH, 'list_loading_service.cfg'))

    @staticmethod
    def _assert_callback(mock_requests_wrapper_post, success, file, error=None):
        data = {
            'success': success,
            'file': file,
            'links': {
                'self': {
                    'href': 'url'
                }
            }
        }
        if success:
            data['links']['member'] = {'href': '/service/id/{member-id}'}

        mock_requests_wrapper_post.assert_has_calls([
            mock.call(url='callback',
                      headers={
                          'PTS-LCP-Base-URL': 'http://localhost',
                          'PTS-LCP-Mode': context.MODE_LIVE,
                          'PTS-LCP-CID': CORRELATION_ID,
                          'PTS-LCP-Principal': PRINCIPAL,
                          'Content-Type': 'application/json'
                      },
                      data=json.dumps(data)
                      )
        ])

    @mock.patch.object(os, 'stat', autospec=True)
    @mock.patch.object(os, 'rename', autospec=True)
    @mock.patch.object(elastic.requests_wrapper, 'post', autospec=True)
    @mock.patch.object(helpers, 'bulk', autospec=True)
    @mock.patch.object(elastic, 'ElasticSearchClient', autospec=True)
    @mock.patch.object(csv, 'reader', autospec=True)
    @mock.patch.object(services.elastic, 'open', create=True)
    @mock.patch.object(os.path, 'isfile', autospec=True)
    def test_create_list_with_csv(self, mock_is_file, mock_open, mock_csv_reader, mock_elastic_search,
                                  mock_bulk, mock_requests_wrapper_post, mock_os_rename, mock_stat):

        mock_open.return_value = mock.MagicMock(spec=file)
        mock_csv_reader.return_value = CsvMock([['abc']])
        mock_elastic_search.return_value = mock.MagicMock()
        mock_elastic_search.return_value.indices.exists.return_value = False
        mock_requests_wrapper_post.return_value = MockHttpResponse(httplib.OK, {})
        mock_stat.return_value = mock.MagicMock()
        mock_stat.return_value.st_size = 1
        request = models.Request(**self.data)

        self.service.create_list(request)

        mock_open.assert_called_with('/content/list_upload/id.csv', 'rU')
        args, _ = mock_bulk.call_args
        tools.assert_equals(type(args[1]), types.GeneratorType)
        mock_bulk.assert_called_with(mock_elastic_search.return_value,
                                     mock.ANY,
                                     index='service',
                                     doc_type='id')
        mock_elastic_search.return_value.indices.refresh.assert_called_once_with(index='service')
        self._assert_callback(mock_requests_wrapper_post, True, 'id.csv')

    @mock.patch.object(os, 'rename', autospec=True)
    @mock.patch.object(elastic.requests_wrapper, 'post', autospec=True)
    @mock.patch.object(helpers, 'bulk', autospec=True)
    @mock.patch.object(configuration, 'data', autospec=True)
    @mock.patch.object(elastic, 'ElasticSearchClient', autospec=True)
    @mock.patch.object(openpyxl, 'load_workbook', autospec=True)
    @mock.patch.object(os.path, 'isfile', autospec=True)
    def test_elastic_search_operation_excel(self, mock_is_file, mock_load_workbook, mock_elastic_search, mock_config,
                                            mock_bulk, mock_requests_wrapper_post, mock_os_rename):
        mock_load_workbook.return_value.active.rows = [[MockCell('abc')]]

        mock_elastic_search.return_value = mock.MagicMock()
        mock_requests_wrapper_post.return_value = MockHttpResponse(httplib.OK, {})
        data = copy.deepcopy(self.data)
        data['filePath'] = 'file.xlsx'
        request = models.Request(**data)
        self.service.create_list(request)

        args, _ = mock_bulk.call_args
        tools.assert_equals(type(args[1]), types.GeneratorType)
        mock_bulk.assert_called_with(mock_elastic_search.return_value,
                                     mock.ANY,
                                     index='service',
                                     doc_type='id')
        mock_elastic_search.return_value.indices.refresh.assert_called_with(index='service')
        self._assert_callback(mock_requests_wrapper_post, True, 'id.xlsx')

    @mock.patch.object(elastic.requests_wrapper, 'post', autospec=True)
    def test_elastic_search_operation_without_callback(self, mock_requests_wrapper_post):
        self.data.pop('callbackUrl')
        mock_requests_wrapper_post.return_value = MockHttpResponse(httplib.OK, {})
        request = models.Request(**self.data)
        self.service.create_list(request)
        mock_requests_wrapper_post.assert_has_calls([])

    @mock.patch.object(os, 'stat', autospec=True)
    @mock.patch.object(os, 'rename', autospec=True)
    @mock.patch.object(elastic.requests_wrapper, 'post', autospec=True)
    @mock.patch.object(elastic, 'ElasticSearchClient', autospec=True)
    @mock.patch.object(csv, 'reader', autospec=True)
    @mock.patch.object(services.elastic, 'open', create=True)
    @mock.patch.object(os.path, 'isfile', autospec=True)
    def test_create_throws_error_on_bulk_call(self, mock_is_file, mock_open, mock_csv_reader, mock_elastic_search,
                                              mock_requests_wrapper_post, mock_os_rename, mock_stat):

        mock_serializer = mock.MagicMock()
        mock_serializer.dumps = lambda *args: json.dumps(args[0])
        mock_open.return_value = mock.MagicMock(spec=file)
        mock_csv_reader.return_value = CsvMock([['abc']])
        mock_stat.return_value.st_size = 1
        mock_elastic_search.return_value = mock.MagicMock()
        mock_elastic_search.return_value.transport = mock.MagicMock()
        mock_elastic_search.return_value.transport.serializer = mock_serializer
        mock_elastic_search.return_value.bulk.side_effect = INTERNAL_SERVER_ERROR_EXCEPTION
        mock_requests_wrapper_post.return_value = MockHttpResponse(httplib.OK, {})
        request = models.Request(**self.data)

        self.service.create_list(request)

        tools.assert_equal(0, mock_elastic_search.return_value.indices.refresh.call_count)
        mock_elastic_search.return_value.bulk.assert_called_once_with(
            "\n".join([
                json.dumps({'index': {'_type': 'id', '_id': 'abc', '_index': 'service'}}),
                json.dumps({'accountNumber': 'abc'})]) + "\n",
            index='service',
            doc_type='id')
        tools.assert_equal(1, mock_elastic_search.return_value.bulk.call_count)
        tools.assert_equal(0, mock_elastic_search.return_value.indices.put_mapping.call_count)
        tools.assert_equal(0, mock_elastic_search.return_value.indices.refresh.call_count)
        self._assert_callback(mock_requests_wrapper_post, False, 'id.csv', "TransportError(500, 'Server error')")

    @mock.patch.object(elastic.logger, 'error', autospec=True)
    @mock.patch.object(elastic.requests_wrapper, 'post', autospec=True)
    @mock.patch.object(os.path, 'isfile', autospec=True)
    def test_create_list_logs_and_returns_error_on_non_existent_csv_file(
            self, mock_is_file, mock_requests_wrapper_post, mock_logger):
        mock_is_file.return_value = False
        request = models.Request(**self.data)

        self.service.create_list(request)
        mock_logger.assert_has_calls(mock.call('An error occurred when creating a new list: File '
                                               '/content/list_upload/file.csv does not exist!'))
        self._assert_callback(mock_requests_wrapper_post, False, 'id.csv',
                              "File /content/list_upload/file.csv does not exist!")

    @mock.patch.object(elastic.logger, 'error', autospec=True)
    @mock.patch.object(os, 'stat', autospec=True)
    @mock.patch.object(os, 'rename', autospec=True)
    @mock.patch.object(elastic.requests_wrapper, 'post', autospec=True)
    @mock.patch.object(csv, 'reader', autospec=True)
    @mock.patch.object(services.elastic, 'open', create=True)
    @mock.patch.object(os.path, 'isfile', autospec=True)
    def test_create_list_logs_and_returns_error_on_empty_csv_file(
            self, mock_is_file, mock_open, mock_csv_reader, mock_requests_wrapper_post, mock_os_rename, mock_stat,
            mock_logger):
        mock_open.return_value = mock.MagicMock(spec=file)
        mock_csv_reader.return_value = CsvMock([['abc']])
        mock_stat.return_value.st_size = 0
        request = models.Request(**self.data)

        self.service.create_list(request)
        mock_logger.assert_has_calls(mock.call('An error occurred when creating a new list: File '
                                               '/content/list_upload/file.csv is empty!'))
        self._assert_callback(mock_requests_wrapper_post, False, 'id.csv',
                              "File /content/list_upload/file.csv is empty!")

    @mock.patch.object(elastic.logger, 'error', autospec=True)
    @mock.patch.object(os, 'rename', autospec=True)
    @mock.patch.object(elastic.requests_wrapper, 'post', autospec=True)
    @mock.patch.object(openpyxl, 'load_workbook', autospec=True)
    @mock.patch.object(services.elastic, 'open', create=True)
    @mock.patch.object(os.path, 'isfile', autospec=True)
    def test_create_list_logs_and_returns_error_on_empty_xlsx_file(
            self, mock_is_file, mock_open, mock_xlsx_reader, mock_requests_wrapper_post, mock_os_rename, mock_logger):
        mock_open.return_value = mock.MagicMock(spec=file)
        mock_xlsx_reader.return_value.active.rows = []
        data = copy.deepcopy(self.data)
        data['filePath'] = 'file.xlsx'
        request = models.Request(**data)

        self.service.create_list(request)
        mock_logger.assert_has_calls(mock.call('An error occurred when creating a new list: File '
                                               '/content/list_upload/file.xlsx is empty!'))
        self._assert_callback(mock_requests_wrapper_post, False, 'id.xlsx',
                              "File /content/list_upload/file.xlsx is empty!")

    @mock.patch.object(os, 'remove')
    @mock.patch.object(elastic, 'ElasticSearchClient', autospec=True)
    def test_delete_list(self, mock_elastic_search, mock_remove):
        mock_elastic_search.return_value = mock.MagicMock()
        request = models.Request(**self.data)
        self.service.delete_list(request)
        mock_elastic_search.return_value.indices.delete_mapping.assert_called_once_with(doc_type='id',
                                                                                        index='service')
        file_path = os.path.join(configuration.data.VOLUME_MAPPINGS_FILE_UPLOAD_TARGET, request.filePath)
        mock_remove.assert_called_once_with(file_path)

    @tools.raises(LookupError)
    @mock.patch.object(elastic, 'ElasticSearchClient', autospec=True)
    def test_delete_list_not_found(self, mock_elastic_search):
        mock_elastic_search.return_value = mock.MagicMock()
        mock_elastic_search.return_value.indices.delete_mapping.side_effect = NOT_FOUND_EXCEPTION
        request = models.Request(**self.data)
        self.service.delete_list(request)

    @tools.raises(exceptions.TransportError)
    @mock.patch.object(elastic, 'ElasticSearchClient', autospec=True)
    def test_delete_list_general_error(self, mock_elastic_search):
        mock_elastic_search.return_value = mock.MagicMock()
        mock_elastic_search.return_value.indices.delete_mapping.side_effect = INTERNAL_SERVER_ERROR_EXCEPTION
        request = models.Request(**self.data)
        self.service.delete_list(request)

    @tools.raises(Exception)
    @mock.patch.object(elastic, 'ElasticSearchClient', autospec=True)
    def test_delete_list_not_acknowledged(self, mock_elastic_search):
        mock_elastic_search.return_value = mock.MagicMock()
        mock_elastic_search.return_value.indices.delete_mapping.return_value = (
            builders.ESDeleteResponseBuilder().with_unacknowledged_response().http_response()['response'])
        request = models.Request(**self.data)
        self.service.delete_list(request)

    @mock.patch.object(os, 'remove')
    @mock.patch.object(elastic, 'ElasticSearchClient', autospec=True)
    def test_delete_list_os_raises_exception(self, mock_elastic_search, mock_remove):
        mock_elastic_search.return_value = mock.MagicMock()
        request = models.Request(**self.data)
        mock_remove.side_effect = OSError

        self.service.delete_list(request)

        mock_elastic_search.return_value.indices.delete_mapping.assert_called_once_with(
            doc_type='id', index='service')
        file_path = os.path.join(configuration.data.VOLUME_MAPPINGS_FILE_UPLOAD_TARGET, request.filePath)
        mock_remove.assert_called_once_with(file_path)

    @mock.patch.object(elastic, 'ElasticSearchClient', autospec=True)
    def test_list_status(self, mock_elastic_search):
        request = models.Request(**self.data)
        response = self.service.get_list_status(request)
        tools.assert_equal(mock_elastic_search.return_value.search.return_value, response)
        mock_elastic_search.return_value.search.assert_called_once_with(
            doc_type='id', index='service', search_type='count')

    @tools.raises(LookupError)
    @mock.patch.object(elastic, 'ElasticSearchClient', autospec=True)
    def test_list_status_throws_404_when_count_0(self, mock_elastic_search):
        mock_elastic_search.return_value.search.return_value = {"hits": {"total": 0}}
        request = models.Request(**self.data)
        self.service.get_list_status(request)

    @mock.patch.object(elastic, 'ElasticSearchClient', autospec=True)
    def test_list_member(self, mock_elastic_search):
        mock_elastic_search.return_value.exists.return_value = True
        request = models.Request(**self.member_data)
        response = self.service.get_list_member(request)
        tools.assert_equal({}, response)
        mock_elastic_search.return_value.exists.assert_called_once_with(doc_type='id', index='service', id='member_id')

    @tools.raises(LookupError)
    @mock.patch.object(elastic, 'ElasticSearchClient', autospec=True)
    def test_list_member_not_found(self, mock_elastic_search):
        mock_elastic_search.return_value.exists.return_value = False
        request = models.Request(**self.member_data)
        response = self.service.get_list_member(request)
        tools.assert_equal({}, response)
        mock_elastic_search.return_value.exists.assert_called_once_with(doc_type='id', index='service', id='member_id')


class TestElasticSearchClient(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        configuration.configure_from(os.path.join(configuration.CONFIGURATION_PATH, 'list_loading_service.cfg'))

    def setUp(self):
        self.client = elastic.ElasticSearchClient()

    @tools.raises(exceptions.TransportError)
    @mock.patch.object(elastic.elasticsearch.client, 'Elasticsearch', autospec=True)
    def test_raises_exception_when_creating_index(self, mock_elastic_search):
        self.client.indices = mock.MagicMock()
        self.client.indices.create.side_effect = INTERNAL_SERVER_ERROR_EXCEPTION

        self.client.bulk([{'_type': 'id', '_id': 'abc', '_source': {'accountNumber': 'abc'}, '_index': 'service'}],
                         index='service',
                         doc_type='id')

        tools.assert_equal(0, self.client.indices.put_mapping.call_count)
        tools.assert_equal(0, self.client.indices.refresh.call_count)
        tools.assert_equal(0, self.client.bulk.call_count)

    @tools.raises(exceptions.TransportError)
    @mock.patch.object(elastic.elasticsearch.client, 'Elasticsearch', autospec=True)
    def test_raises_exception_when_mapping_put_raises_exception(self, mock_elastic_search):
        self.client.indices = mock.MagicMock()
        self.client.indices.exists.side_effect = INTERNAL_SERVER_ERROR_EXCEPTION

        self.client.bulk([{'_type': 'id', '_id': 'abc', '_source': {'accountNumber': 'abc'}, '_index': 'service'}],
                         index='service',
                         doc_type='id')

        self.client.indices.exists.assert_called_with_once(index='service')
        tools.assert_equal(0, self.client.indices.create.call_count)
        tools.assert_equal(0, self.client.indices.refresh.call_count)
        tools.assert_equal(0, self.client.indices.put_mapping.call_count)

    @tools.raises(exceptions.TransportError)
    @mock.patch.object(elastic.elasticsearch.client, 'Elasticsearch', autospec=True)
    def test_raises_exception_when_mapping_put(self, mock_elastic_search):
        self.client.indices = mock.MagicMock()
        self.client.indices.exists.return_value = True
        self.client.indices.put_mapping.side_effect = INTERNAL_SERVER_ERROR_EXCEPTION

        self.client.bulk([{'_type': 'id', '_id': 'abc', '_source': {'accountNumber': 'abc'}, '_index': 'service'}],
                         index='service',
                         doc_type='id')

        self.client.indices.exists.assert_called_with_once(index='service')
        tools.assert_equal(1, self.client.indices.create.call_count)
        tools.assert_equal(1, self.client.indices.refresh.call_count)
        self.client.indices.put_mapping.assert_called_once_with(
            index='service',
            doc_type='id',
            body={
                "properties": {
                    "accountNumber": {
                        "type": "string",
                        "index": "not_analyzed"
                    }
                }
            })

    @mock.patch.object(elastic.elasticsearch.client.Elasticsearch, 'bulk', autospec=True)
    def test_create_index_and_mapping_success(self, mock_elastic_search_bulk_method):
        self.client.indices = mock.MagicMock()
        self.client.indices.exists.return_value = False

        self.client.bulk([{'_type': 'id', '_id': 'abc', '_source': {'accountNumber': 'abc'}, '_index': 'service'}],
                         index='service',
                         doc_type='id')

        self.client.indices.exists.assert_called_once_with(index='service')
        self.client.indices.create.assert_called_once_with(index='service')
        self.client.indices.put_mapping.assert_called_once_with(index='service',
                                                                doc_type='id',
                                                                body={
                                                                    "properties": {
                                                                        "accountNumber": {
                                                                            "type": "string",
                                                                            "index": "not_analyzed"
                                                                        }
                                                                    }
                                                                })
        mock_elastic_search_bulk_method.assert_called_once_with(
            self.client,
            [{'_type': 'id', '_id': 'abc', '_source': {'accountNumber': 'abc'}, '_index': 'service'}],
            index='service',
            params={},
            doc_type='id'
        )
