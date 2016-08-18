import csv
import os
import types
import unittest

import mock
import openpyxl
from nose import tools

import configuration
from app.services import readers
from tests import mocks


class TestElasticSearchFileReaders(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        configuration.configure_from(os.path.join(configuration.CONFIGURATION_PATH, 'list_loading_service.cfg'))

    @tools.raises(Exception)
    def test_should_always_be_a_instance_of_csv_xls_reader(self):
        readers.FileReader('/content/list_upload/id.csv')

    @mock.patch.object(csv, 'reader', autospec=True)
    @mock.patch.object(os.path, 'isfile', autospec=True)
    @mock.patch.object(readers, 'open', create=True)
    def test_get_correct_reader_on_extension_csv(self, mock_open, mock_is_file, mock_csv_reader):
        mock_account_no = mock.MagicMock(return_value='account_no')
        account_no_list = [mock_account_no]
        mock_open.return_value = iter(account_no_list)
        mock_csv_reader.return_value.get_rows.return_value = mocks.generator(account_no_list)

        tools.assert_true(
            isinstance(readers.BulkAccountsFileReaders.get('/content/list_upload/id.csv'), readers.CsvReader))

    @mock.patch.object(csv, 'reader', autospec=True)
    @mock.patch.object(os.path, 'isfile', autospec=True)
    @mock.patch.object(readers, 'open', create=True)
    def test_get_correct_reader_on_extension_txt(self, mock_open, mock_is_file, mock_csv_reader):
        mock_account_no = mock.MagicMock(return_value='account_no')
        account_no_list = [mock_account_no]
        mock_open.return_value = iter(account_no_list)

        mock_csv_reader.return_value = mocks.generator(account_no_list)
        tools.assert_true(
            isinstance(readers.BulkAccountsFileReaders.get('/content/list_upload/id.txt'), readers.CsvReader))

    @mock.patch.object(os.path, 'isfile', autospec=True)
    @mock.patch.object(csv, 'reader', autospec=True)
    @mock.patch.object(readers, 'open', create=True)
    def test_reads_csv_file_correctly(self, mock_open, mock_csv_reader, mock_is_file):
        accounts = ['account_no']
        mock_open.return_value = iter(accounts)
        mock_csv_reader.return_value = mocks.generator(accounts)

        csv_reader = readers.CsvReader('/content/list_upload/id.csv')
        tools.assert_equals(mocks.Any(types.GeneratorType), csv_reader.get_rows())
        tools.assert_equals("account_no", next(csv_reader.get_rows()))
        try:
            next(next(csv_reader.get_rows()))
        except StopIteration as e:
            tools.assert_is_not_none(e)

    @mock.patch.object(os.path, 'isfile', autospec=True)
    @mock.patch.object(csv, 'reader', autospec=True)
    @mock.patch.object(readers, 'open', create=True)
    def test_reads_csv_file_and_return_if_exceeds_allowed_count(self, mock_open, mock_csv_reader, mock_is_file):
        configuration.data.ACCOUNTS_UPDATE_MAX_SIZE_ALLOWED = 1
        accounts = ['account_no' for _ in xrange(2)]
        mock_open.return_value = iter(accounts)
        mock_csv_reader.return_value = mocks.generator(accounts)

        tools.assert_true(readers.CsvReader('/content/list_upload/id.csv').exceeds_allowed_row_count())

    @mock.patch.object(os.path, 'isfile', autospec=True)
    @mock.patch.object(openpyxl, 'load_workbook', autospec=True)
    def test_reads_excel_file_correctly(self, mock_load_workbook, mock_is_file):
        mock_cell = mock.MagicMock()
        mock_cell.value = 'account_no'
        mock_load_workbook.return_value.active.rows = mocks.generator([mock_cell])

        xl_reader = readers.ExcelReader('/content/list_upload/id.xlsx')

        mock_load_workbook.return_value.active.rows = mocks.generator([mock_cell])
        tools.assert_equals(mocks.Any(types.GeneratorType), xl_reader.get_rows())
        tools.assert_equals("account_no", next(xl_reader.get_rows()))
        tools.assert_is_none(xl_reader.close())

    @mock.patch.object(os.path, 'isfile', autospec=True)
    @mock.patch.object(openpyxl, 'load_workbook', autospec=True)
    def test_reads_xlsx_file_and_return_if_exceeds_allowed_count(self, mock_load_workbook, mock_is_file):
        configuration.data.ACCOUNTS_UPDATE_MAX_SIZE_ALLOWED = 1
        mock_cell = mock.MagicMock()
        mock_cell.value = 'account_no'
        mock_load_workbook.return_value.active.rows = mocks.generator([mock_cell for _ in xrange(2)])

        tools.assert_true(readers.ExcelReader('/content/list_upload/id.xlsx').exceeds_allowed_row_count())

    @mock.patch.object(os.path, 'isfile', autospec=True)
    @mock.patch.object(openpyxl, 'load_workbook', autospec=True)
    def test_get_correct_reader_on_extension_xlsx(self, mock_load_workbook, mock_is_file):
        mock_cell = mock.MagicMock()
        mock_cell.value = 'account_no'
        account_no_list = [mock_cell]

        mock_load_workbook.return_value.active.rows = mocks.generator(account_no_list)
        tools.assert_true(
            isinstance(readers.BulkAccountsFileReaders.get('/content/list_upload/id.xlsx'), readers.ExcelReader))
